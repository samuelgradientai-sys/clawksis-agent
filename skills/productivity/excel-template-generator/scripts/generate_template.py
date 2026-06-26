#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except Exception as exc:
    print("ERROR: falta openpyxl.", file=sys.stderr)
    print("Instala con:", file=sys.stderr)
    print(
        '/opt/clawksis-agent/.venv/bin/python -m pip install "openpyxl==3.1.5"',
        file=sys.stderr,
    )
    print(f"Detalle: {exc!r}", file=sys.stderr)
    raise SystemExit(2)


HEADER_FILL = "1F4E79"
SUBHEADER_FILL = "D9EAF7"
LOW_FILL = "FFC7CE"
BORDER_COLOR = "D9E2F3"


def safe_output_path(raw: str | None, template: str) -> Path:
    if raw:
        path = Path(raw).expanduser()
    else:
        out_dir = Path.home() / "clawksis_exports" / "excel_templates"
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        path = out_dir / f"{template}_{stamp}.xlsx"

    if path.suffix.lower() != ".xlsx":
        raise SystemExit("ERROR: la salida debe terminar en .xlsx")

    path.parent.mkdir(parents=True, exist_ok=True)
    return path.resolve()


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_sheet(ws) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def create_info_sheet(wb: Workbook, business_name: str, template: str) -> None:
    ws = wb.active
    ws.title = "Instrucciones"

    ws["A1"] = f"Plantilla Excel — {business_name}"
    ws["A2"] = f"Tipo: {template}"
    ws["A3"] = f"Generada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A5"] = "Uso"
    ws["B5"] = (
        "Completa las filas editables. Las columnas con fórmulas se calculan automáticamente."
    )
    ws["A6"] = "Seguridad"
    ws["B6"] = "Este archivo no contiene macros ni conexiones externas."
    ws["A7"] = "Recomendación"
    ws["B7"] = "Conserva una copia original antes de usarla en operación real."

    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")

    for cell in ["A5", "A6", "A7"]:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill("solid", fgColor=SUBHEADER_FILL)

    set_widths(ws, {"A": 22, "B": 80})


def create_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Listas")

    data = {
        "A": ["Categorías", "Lentes", "Monturas", "Accesorios", "Servicios", "Otro"],
        "B": ["Estados", "Activo", "Inactivo", "Agotado", "Pendiente", "Cerrado"],
        "C": [
            "Métodos de pago",
            "Efectivo",
            "Transferencia",
            "Tarjeta",
            "Crédito",
            "Otro",
        ],
        "D": ["Canales", "Tienda", "WhatsApp", "Telegram", "Referido", "Otro"],
        "E": ["Prioridades", "Baja", "Media", "Alta", "Crítica"],
    }

    for col, values in data.items():
        for i, value in enumerate(values, start=1):
            ws[f"{col}{i}"] = value

    style_sheet(ws)
    set_widths(ws, {"A": 20, "B": 18, "C": 20, "D": 18, "E": 18})
    ws.sheet_state = "hidden"


def inventory_template(wb: Workbook, rows: int) -> None:
    ws = wb.create_sheet("Inventario")

    headers = [
        "SKU",
        "Producto",
        "Categoría",
        "Proveedor",
        "Stock actual",
        "Stock mínimo",
        "Costo unitario COP",
        "Precio venta COP",
        "Margen COP",
        "Margen %",
        "Valor inventario COP",
        "Estado",
        "Notas",
    ]
    ws.append(headers)

    for r in range(2, rows + 2):
        ws[f"I{r}"] = f'=IF(OR(G{r}="",H{r}=""),"",H{r}-G{r})'
        ws[f"J{r}"] = f'=IF(OR(G{r}="",H{r}="",H{r}=0),"",I{r}/H{r})'
        ws[f"K{r}"] = f'=IF(OR(E{r}="",G{r}=""),"",E{r}*G{r})'
        ws[f"L{r}"] = f'=IF(E{r}="","",IF(E{r}<=F{r},"Stock bajo","OK"))'

    style_sheet(ws)
    add_table(ws, "InventarioTable", f"A1:M{rows + 1}")
    set_widths(
        ws,
        {
            "A": 16,
            "B": 28,
            "C": 18,
            "D": 22,
            "E": 14,
            "F": 14,
            "G": 18,
            "H": 18,
            "I": 16,
            "J": 12,
            "K": 20,
            "L": 14,
            "M": 30,
        },
    )

    for col in ["G", "H", "I", "K"]:
        for row in ws[f"{col}2:{col}{rows + 1}"]:
            row[0].number_format = "$ #,##0"

    for row in ws[f"J2:J{rows + 1}"]:
        row[0].number_format = "0.00%"

    dv_cat = DataValidation(type="list", formula1="=Listas!$A$2:$A$6", allow_blank=True)
    dv_status = DataValidation(
        type="list", formula1='"Stock bajo,OK,Agotado,Inactivo"', allow_blank=True
    )

    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_status)
    dv_cat.add(f"C2:C{rows + 1}")
    dv_status.add(f"L2:L{rows + 1}")

    ws.conditional_formatting.add(
        f"L2:L{rows + 1}",
        FormulaRule(
            formula=['L2="Stock bajo"'], fill=PatternFill("solid", fgColor=LOW_FILL)
        ),
    )


def sales_template(wb: Workbook, rows: int) -> None:
    ws = wb.create_sheet("Ventas")

    headers = [
        "Fecha",
        "ID venta",
        "Cliente",
        "Producto/SKU",
        "Cantidad",
        "Precio unitario COP",
        "Descuento COP",
        "Total COP",
        "Método de pago",
        "Canal",
        "Estado",
        "Notas",
    ]
    ws.append(headers)

    for r in range(2, rows + 2):
        ws[f"H{r}"] = f'=IF(OR(E{r}="",F{r}=""),"",E{r}*F{r}-IF(G{r}="",0,G{r}))'

    style_sheet(ws)
    add_table(ws, "VentasTable", f"A1:L{rows + 1}")
    set_widths(
        ws,
        {
            "A": 14,
            "B": 16,
            "C": 26,
            "D": 28,
            "E": 12,
            "F": 18,
            "G": 18,
            "H": 16,
            "I": 18,
            "J": 16,
            "K": 14,
            "L": 30,
        },
    )

    for col in ["F", "G", "H"]:
        for row in ws[f"{col}2:{col}{rows + 1}"]:
            row[0].number_format = "$ #,##0"

    for row in ws[f"A2:A{rows + 1}"]:
        row[0].number_format = "yyyy-mm-dd"

    dv_pay = DataValidation(type="list", formula1="=Listas!$C$2:$C$6", allow_blank=True)
    dv_channel = DataValidation(
        type="list", formula1="=Listas!$D$2:$D$6", allow_blank=True
    )
    dv_status = DataValidation(
        type="list", formula1='"Pendiente,Pagada,Anulada,Devuelta"', allow_blank=True
    )

    ws.add_data_validation(dv_pay)
    ws.add_data_validation(dv_channel)
    ws.add_data_validation(dv_status)

    dv_pay.add(f"I2:I{rows + 1}")
    dv_channel.add(f"J2:J{rows + 1}")
    dv_status.add(f"K2:K{rows + 1}")


def postventa_template(wb: Workbook, rows: int) -> None:
    ws = wb.create_sheet("Postventa")

    headers = [
        "Fecha caso",
        "Cliente",
        "Teléfono",
        "Producto/Servicio",
        "Tipo de caso",
        "Estado",
        "Prioridad",
        "Responsable",
        "Costo adicional COP",
        "Próxima acción",
        "Fecha compromiso",
        "Días abiertos",
        "Notas",
    ]
    ws.append(headers)

    for r in range(2, rows + 2):
        ws[f"L{r}"] = f'=IF(A{r}="","",TODAY()-A{r})'

    style_sheet(ws)
    add_table(ws, "PostventaTable", f"A1:M{rows + 1}")
    set_widths(
        ws,
        {
            "A": 14,
            "B": 26,
            "C": 16,
            "D": 26,
            "E": 20,
            "F": 16,
            "G": 14,
            "H": 18,
            "I": 18,
            "J": 30,
            "K": 16,
            "L": 14,
            "M": 34,
        },
    )

    for row in ws[f"I2:I{rows + 1}"]:
        row[0].number_format = "$ #,##0"

    for col in ["A", "K"]:
        for row in ws[f"{col}2:{col}{rows + 1}"]:
            row[0].number_format = "yyyy-mm-dd"

    dv_state = DataValidation(
        type="list",
        formula1='"Abierto,En proceso,Esperando cliente,Cerrado,Cancelado"',
        allow_blank=True,
    )
    dv_prio = DataValidation(
        type="list", formula1="=Listas!$E$2:$E$5", allow_blank=True
    )

    ws.add_data_validation(dv_state)
    ws.add_data_validation(dv_prio)

    dv_state.add(f"F2:F{rows + 1}")
    dv_prio.add(f"G2:G{rows + 1}")

    ws.conditional_formatting.add(
        f"G2:G{rows + 1}",
        FormulaRule(
            formula=['G2="Crítica"'], fill=PatternFill("solid", fgColor=LOW_FILL)
        ),
    )


def create_summary(wb: Workbook, template: str, rows: int) -> None:
    ws = wb.create_sheet("Resumen")

    ws["A1"] = "Resumen"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")

    if template == "inventory":
        items = [
            ("Total unidades", f"=SUM(Inventario!E2:E{rows + 1})"),
            ("Valor inventario COP", f"=SUM(Inventario!K2:K{rows + 1})"),
            (
                "Productos con stock bajo",
                f'=COUNTIF(Inventario!L2:L{rows + 1},"Stock bajo")',
            ),
        ]
    elif template == "sales":
        items = [
            ("Ventas registradas", f"=COUNTA(Ventas!B2:B{rows + 1})"),
            ("Unidades vendidas", f"=SUM(Ventas!E2:E{rows + 1})"),
            ("Total vendido COP", f"=SUM(Ventas!H2:H{rows + 1})"),
        ]
    else:
        items = [
            ("Casos abiertos", f'=COUNTIF(Postventa!F2:F{rows + 1},"Abierto")'),
            ("Casos en proceso", f'=COUNTIF(Postventa!F2:F{rows + 1},"En proceso")'),
            ("Costos adicionales COP", f"=SUM(Postventa!I2:I{rows + 1})"),
        ]

    ws.append(["Métrica", "Valor"])
    for label, formula in items:
        ws.append([label, formula])

    style_sheet(ws)
    set_widths(ws, {"A": 32, "B": 24})

    for cell in ws["B"]:
        cell.number_format = "$ #,##0"


def build_workbook(template: str, business_name: str, rows: int) -> Workbook:
    wb = Workbook()
    create_info_sheet(wb, business_name, template)
    create_lists_sheet(wb)

    if template == "inventory":
        inventory_template(wb, rows)
    elif template == "sales":
        sales_template(wb, rows)
    elif template == "postventa":
        postventa_template(wb, rows)
    else:
        raise SystemExit(f"ERROR: template no soportado: {template}")

    create_summary(wb, template, rows)
    return wb


def validate_generated_file(path: Path) -> None:
    wb = load_workbook(path, data_only=False)

    required = {"Instrucciones", "Listas", "Resumen"}
    missing = sorted(required - set(wb.sheetnames))
    if missing:
        raise SystemExit(f"ERROR: faltan hojas requeridas: {missing}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate safe Excel .xlsx templates for Clawksis operations."
    )
    parser.add_argument(
        "--template", choices=["inventory", "sales", "postventa"], required=True
    )
    parser.add_argument("--business-name", default="Clawksis")
    parser.add_argument("--rows", type=int, default=200)
    parser.add_argument("--output", default=None)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    if args.rows < 10 or args.rows > 5000:
        raise SystemExit("ERROR: --rows debe estar entre 10 y 5000")

    output = safe_output_path(args.output, args.template)

    if output.exists() and not args.overwrite:
        raise SystemExit(
            f"ERROR: el archivo ya existe. Usa --overwrite si confirmas sobrescribir: {output}"
        )

    wb = build_workbook(args.template, args.business_name, args.rows)
    wb.save(output)
    validate_generated_file(output)

    download_url = "/artifacts/download?path=" + quote(str(output), safe="")
    print(f"OK: archivo generado: {output}")
    print(f"template={args.template}")
    print("macros=none")
    print(f"download_url={download_url}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
